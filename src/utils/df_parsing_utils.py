import pandas as pd
import os
from typing import List
from openpyxl import load_workbook
import numpy as np

from src.utils.logger import logger


 ################################################# Common file interactions #################################################
def _xlsx_sheets_to_csvs(file_path: str, header_row: int = None) -> None:
    """
    Read all sheets of an XLSX file, clean, and write each as a CSV named filebase_sheetname.csv
    in the same folder as the XLSX file.
    Raises ValueError if no sheets or all sheets are empty.
    """
    try:
        sheet_names = load_workbook(file_path, read_only=True).sheetnames
        if not sheet_names:
            raise ValueError(f"No sheets found in {file_path}")
        base_name = os.path.splitext(os.path.basename(file_path))[0]
        output_folder = os.path.dirname(file_path)
        written = 0
        for sheet in sheet_names:
            if header_row is not None:
                df = pd.read_excel(file_path, sheet_name=sheet,header = header_row, engine='openpyxl')
            else:
                df = pd.read_excel(file_path, sheet_name=sheet, engine='openpyxl')
            df = _apply_common_rules(df)
            if not df.empty:
                safe_sheet = str(sheet).replace(" ", "_").replace("/", "_")
                output_path = os.path.join(output_folder, f"{base_name}_{safe_sheet}.csv")
                df.to_csv(output_path, index=False)
                logger.info(f"Saved sheet '{sheet}' to {output_path}")
                written += 1
        if written == 0:
            raise ValueError(f"All sheets in {file_path} are empty after cleaning.")
    except Exception as e:
        logger.error(f"Failed to process XLSX file {file_path}: {e}")
        raise


def _construct_file_paths(folder: str) -> List[str]:
    """Construct a full file path for all files in the folder."""
    file_paths = []
    files = [f for f in os.listdir(folder) if os.path.isfile(os.path.join(folder, f))]
    for f in files:
        file_path = os.path.join(folder, f)
        file_paths.append(file_path)
    return file_paths





def convert_xls_to_xlsx(folder: str, output_folder: str = None):
    """
    Convert all .xls files in a folder to .xlsx format.
    Args:
        folder: Source folder containing .xls files.
        output_folder: Destination folder for .xlsx files (defaults to source folder).
    """
    if output_folder is None:
        output_folder = folder

    for filename in os.listdir(folder):
        if filename.lower().endswith('.xls') and not filename.lower().endswith('.xlsx'):
            xls_path = os.path.join(folder, filename)
            xlsx_filename = os.path.splitext(filename)[0] + '.xlsx'
            xlsx_path = os.path.join(output_folder, xlsx_filename)
            try:
                # Read all sheets using xlrd
                xls = pd.ExcelFile(xls_path, engine='xlrd')
                with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
                    for sheet_name in xls.sheet_names:
                        df = xls.parse(sheet_name)
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                print(f"Converted {xls_path} -> {xlsx_path}")
            except Exception as e:
                print(f"Failed to convert {xls_path}: {e}")


def delete_xls_files(folder: str):
    files = [f for f in os.listdir(folder) 
             if os.path.isfile(os.path.join(folder, f)) 
             and f.lower().endswith('.xlsx') 
             and 'vacs01' in f.lower()
             and '2017' not in f.lower()
             ]
    for f in files:
        file_path = os.path.join(folder, f)
        try:
            os.remove(file_path)
            logger.info(f"Deleted file: {file_path}")
        except Exception as e:
            logger.error(f"Failed to delete {file_path}: {e}")

################################################# Common rules for formatting dataframes #################################################

def _clean_column_names(cols: list[str]) -> list[str]:
    """Standardize column names to lowercase, snake_case, no spaces."""
    logger.info(f"Cleaning columns: {cols}")
    cleaned = [
        c.strip().lower().replace(" ", "_").replace("-", "_")
        for c in cols
    ]
    logger.debug(f"Cleaned columns: {cleaned}")
    return cleaned

def _apply_common_rules(df: pd.DataFrame) -> pd.DataFrame:
    """Remove rows that are completely empty or contain only placeholders.
    Rules Applied:
    1. Remove rows that are completely empty or contain only NaN values.
    2. Remove columns that are completely empty or contain only NaN values.
    3. Reset the DataFrame index after row removals.
    4. Clean column names to be lowercase and snake_case.
    5. Remove all rows above the row containing the marker "united kingdom (thousands), seasonally adjusted".
    6. Trim whitespace from string entries in the DataFrame.
    7. Replace common placeholders for missing values (e.g., '', 'n/a', '-', '--') with NaN.
    8. Remove duplicate rows.
    9. Convert columns with numeric data stored as strings to appropriate numeric types.
    """
    logger.debug(f"DF shape before common rules applied: {df.shape}")
    # Rule 1
    df.dropna(how='all', inplace=True)
    # Rule 2
    df.dropna(axis=1, how='all', inplace=True)
    # Rule 3
    df.reset_index(drop=True, inplace=True)
    # Rule 4
    df.columns = _clean_column_names(df.columns)
    # Rule 5
    marker = "united kingdom (thousands), seasonally adjusted"
    marker_row = None
    """
    for i in range(df.shape[0]):
        if any(marker in str(val).lower() for val in df.iloc[i]):
            marker_row = i
            break
    if marker_row is not None:
        df = df.iloc[marker_row+1:].reset_index(drop=True)
    """
    # Rule 6
    df = df.apply(lambda col: col.str.strip() if col.dtype == "object" else col)
    # Rule 7
    missing_values = ['', 'n/a', 'na', '-', '--']
    df.replace(missing_values, np.nan, inplace=True)
    # Rule 8 
    df = df.drop_duplicates()
    # Rule 9
    for col in df.select_dtypes(include='object').columns:
       try:
           df[col] = pd.to_numeric(df[col])
       except Exception:
           pass
    # Rule 10
    if 'unnamed: 0' in df.columns:
       df.drop(columns=['unnamed: 0'], inplace=True)
    logger.debug(f"DF shape after common rules applied: {df.shape}")

    return df

