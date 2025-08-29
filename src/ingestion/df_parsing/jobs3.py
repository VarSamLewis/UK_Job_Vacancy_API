from pathlib import Path
#from src.utils.df_parsing_utils  import _xlsx_sheets_to_csvs
import pandas as pd
import os
from typing import List
from openpyxl import load_workbook
import numpy as np

from src.utils.logger import logger

def _xlsx_sheets_to_csvs_jobs03(file_path: str, header_row: int = None) -> None:
    """
    Read all sheets of an XLSX file, clean, and write each as a CSV named filebase_sheetname.csv
    in the same folder as the XLSX file.
    Raises ValueError if no sheets or all sheets are empty.
    """
    try:
        sheet_names = load_workbook(file_path, read_only=True).sheetnames
        if not sheet_names:
            raise ValueError(f"No sheets found in {file_path}")
        base_name = os.path.splitext(os.path.basename(file_path))[0]
        output_folder = os.path.dirname(file_path)
        written = 0
        for sheet in sheet_names:
            df = pd.read_excel(file_path, sheet_name=sheet,header = header_row, engine='openpyxl')
            

            # Drop rows where all elements are NaN
            df.dropna(how='all', inplace=True)

            # Drop columns where all elements are NaN
            df.dropna(axis=0, how='all', inplace=True)

            # Rename the 1st column
            df.rename(columns={df.columns[0]: "Mon"}, inplace=True)
  
            # Drop column 1 (if needed)
            df.drop(df.columns[1], axis=1, inplace=True)
            # Drop rows 1 and 2
            df = df.iloc[2:]
            # Remove trailing annotations like (r), (p), and extra spaces
            df['Mon'] = df['Mon'].astype(str).str.replace(r"\s*\(.*\)", "", regex=True).str.strip()
            # Convert to datetime and format as MM-YY, invalid values become NaT
            df['Mon'] = pd.to_datetime(df['Mon'], errors='coerce', format='%b %y')
            
            if not df.empty:
                safe_sheet = str(sheet).replace(" ", "_").replace("/", "_")
                output_path = os.path.join(output_folder, f"{base_name}_{safe_sheet}.csv")
                df.to_csv(output_path, index=False)
                logger.info(f"Saved sheet '{sheet}' to {output_path}")
                written += 1
        if written == 0:
            raise ValueError(f"All sheets in {file_path} are empty after cleaning.")
    except Exception as e:
        logger.error(f"Failed to process XLSX file {file_path}: {e}")
        raise

if __name__ == "__main__":
    file = Path(r"C:\Users\samle\Source\Repos\UK_Job_Vacancy_API\Data\jobs3jun2025.xlsx")
    _xlsx_sheets_to_csvs_jobs03(file, 3)
