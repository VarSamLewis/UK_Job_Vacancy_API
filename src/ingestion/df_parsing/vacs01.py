from pathlib import Path
from src.utils.df_parsing_utils  import convert_xls_to_xlsx, delete_xls_files
import pandas as pd
import os
from typing import List
from openpyxl import load_workbook
import re

from src.utils.logger import logger


def _xlsx_sheets_to_csvs_vacs01(file_path: str, output_folder: str , header_row: int = None) -> None:
    """
    Read all sheets of an XLSX file, clean, and write each as a CSV named filebase_sheetname.csv
    in the same folder as the XLSX file.
    Raises ValueError if no sheets or all sheets are empty.
    """
    try:
        sheet_names = load_workbook(file_path, read_only=True).sheetnames
        if not sheet_names:
            raise ValueError(f"No sheets found in {file_path}")
        base_name = os.path.splitext(os.path.basename(file_path))[0]
        written = 0
        for sheet in sheet_names:
            df = pd.read_excel(file_path, sheet_name=sheet,header = header_row, engine='openpyxl')
            

            # Drop rows where all elements are NaN
            df.dropna(how='all', inplace=True)

            # Drop columns where all elements are NaN
            df.dropna(axis=0, how='all', inplace=True)

            # Drop rows where first column is NaN
            df.dropna(subset=[df.columns[0]], inplace=True)

            # Rename the 1st column
            df.rename(columns={df.columns[0]: "Quart"}, inplace=True)
  
            # Drop column 1 (if needed)
            df.drop(df.columns[1], axis=1, inplace=True)

            df['Quart'] = df['Quart'].astype(str).str.replace(r"\s*\(.*\)", "", regex=True).str.strip()

            df['year'] = df['Quart'].str.extract(r'(\d{4})')
            # Split "months" into "start_month" and "end_month"
            df['months'] = df['Quart'].str.extract(r'^([A-Za-z\-]+)')

            month_split = df['months'].str.split('-', expand=True)
            df['start_mon_char'] = month_split[0].str.strip()
            df['end_mon_char'] = month_split[1].str.strip() if month_split.shape[1] > 1 else month_split[0].str.strip()

            # Create datetime columns
            df['start_month_str'] = df['start_mon_char'] + ' ' + df['year']
            df['end_month_str'] = df['end_mon_char'] + ' ' + df['year']

            df['start_mon'] = pd.to_datetime(df['start_month_str'], errors='coerce', format='%b %Y')
            df['end_mon'] = pd.to_datetime(df['end_month_str'], errors='coerce', format='%b %Y') + pd.offsets.MonthEnd(0)

            #df['year'] = df['year'].astype(int)

            df.drop(columns=['months', 'start_month_str', 'end_month_str', 'Quart','Unnamed: 5','Unnamed: 6'], inplace=True)

            # Drop rows 1
            df = df.iloc[1:]

            if not df.empty:
                safe_sheet = str(sheet).replace(" ", "_").replace("/", "_")
                output_path = os.path.join(output_folder, f"{base_name}_{safe_sheet}.csv")
                df.to_csv(output_path, index=False)
                logger.info(f"Saved sheet '{sheet}' to {output_path}")
                written += 1
        if written == 0:
            raise ValueError(f"All sheets in {file_path} are empty after cleaning.")
    except Exception as e:
        logger.error(f"Failed to process XLSX file {file_path}: {e}")
        raise

def main():
    folder = Path(r"C:\Users\samle\Source\Repos\UK_Job_Vacancy_API\Data")
    outpath_folder = Path(r"C:\Users\samle\Source\Repos\UK_Job_Vacancy_API\Data\vacs01")
    files = [f for f in os.listdir(folder) 
             if os.path.isfile(os.path.join(folder, f)) 
             and f.lower().endswith('.xlsx') 
             and 'vacs01' in f.lower()
             and '2017' not in f.lower()
             ]

    convert_xls_to_xlsx(folder)
    delete_xls_files(folder)
    

    for file in files:
        file_path = os.path.join(folder, file)
        if not os.path.exists(file_path):
            logger.error(f"File not found: {file_path}")
            continue
        _xlsx_sheets_to_csvs_vacs01(file_path, outpath_folder, 3)
